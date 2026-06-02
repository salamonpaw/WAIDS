from __future__ import annotations

import io
import os
import re
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
from typing import Annotated, List, Optional

import pandas as pd
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, Response, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from pydantic import BaseModel
from psycopg2.extras import execute_values

from database import (
    get_conn, get_status, get_analysis, init_db,
    get_excluded_firms, add_excluded_firm, remove_excluded_firm,
    get_reps, get_all_firms, assign_firm_to_rep, remove_firm_from_rep,
    add_sales_rep, remove_sales_rep,
    set_device_type_override, get_type_overrides,
    get_payments_for_sn,
    set_device_comment, bulk_set_device_type,
    get_firm_configs, upsert_firm_config, delete_firm_config,
    get_firms_for_export, import_firms_table,
    get_monthly_revenue,
    get_license_fees, upsert_license_fee, delete_license_fee,
    merge_firms, get_merge_history,
    # suspensions
    get_device_suspensions, add_device_suspension, delete_device_suspension,
    get_firm_suspensions, add_firm_suspension, delete_firm_suspension,
    get_active_suspensions, get_all_suspensions_map,
    # firm-rep export/import
    get_firms_with_reps, get_firms_without_reps, import_firm_reps,
    get_rep_first_ids_check,
    VALID_FIRM_TYPES, VALID_CYCLES,
    # auth
    get_or_create_secret_key, create_admin_if_needed,
    verify_user, get_user_by_id,
    list_users, create_user_db, set_user_status, reset_user_password,
    # import sessions
    create_import_session, get_import_sessions, undo_import_session,
    # bulk device update
    bulk_update_devices,
)
from version import APP_VERSION

app = FastAPI(title="Weryfikator Abonamentów API", version=APP_VERSION)

# ── JWT config (key stored in DB, survives restarts) ─────────────────────────
_JWT_ALGO   = "HS256"
_TOKEN_TTL  = 8   # hours
_SECRET_KEY = ""  # filled in startup
_bearer     = HTTPBearer(auto_error=False)


def _make_token(user_id: int) -> str:
    exp = datetime.utcnow() + timedelta(hours=_TOKEN_TTL)
    return jwt.encode({"sub": str(user_id), "exp": exp}, _SECRET_KEY, algorithm=_JWT_ALGO)


def _decode_token(token: str) -> Optional[int]:
    try:
        payload = jwt.decode(token, _SECRET_KEY, algorithms=[_JWT_ALGO])
        sub = payload.get("sub")
        return int(sub) if sub else None
    except (JWTError, ValueError):
        return None


# ── auth dependencies ─────────────────────────────────────────────────────────

def get_auth_user(request: Request) -> dict:
    """Read user from request.state (set by auth middleware)."""
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(401, "Not authenticated")
    return user


def require_admin(user: dict = Depends(get_auth_user)) -> dict:
    if not user.get("is_admin"):
        raise HTTPException(403, "Wymagane uprawnienia administratora")
    return user


# ── auth middleware ───────────────────────────────────────────────────────────
_PUBLIC = {"/auth/login", "/version", "/docs", "/openapi.json", "/redoc", "/changelog"}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in _PUBLIC or path.startswith(("/docs", "/openapi", "/redoc")):
        return await call_next(request)
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    uid = _decode_token(auth[7:])
    if uid is None:
        return JSONResponse({"detail": "Invalid token"}, status_code=401)
    user = get_user_by_id(uid)
    if not user or not user["is_active"]:
        return JSONResponse({"detail": "User inactive"}, status_code=401)
    request.state.user = user
    return await call_next(request)


@app.on_event("startup")
def startup():
    global _SECRET_KEY
    init_db()
    _SECRET_KEY = get_or_create_secret_key()
    # Create first admin from env vars if no admins exist
    email    = os.getenv("ADMIN_EMAIL", "p.salamon@asdsystems.pl")
    password = os.getenv("ADMIN_PASSWORD", "")
    if password:
        created = create_admin_if_needed(email, password)
        if created:
            print(f"[WAIDS] ✓ Admin user created: {email}")


@app.get("/version")
def get_version():
    """Zwraca aktualną wersję aplikacji."""
    return {"version": APP_VERSION}


@app.get("/changelog")
def get_changelog():
    """Zwraca zawartość CHANGELOG.md."""
    p = Path(__file__).parent / "CHANGELOG.md"
    return {"content": p.read_text(encoding="utf-8") if p.exists() else "Brak CHANGELOG.md"}


# ── auth endpoints ────────────────────────────────────────────────────────────

class LoginIn(BaseModel):
    email: str
    password: str


@app.post("/auth/login")
def login(body: LoginIn):
    user = verify_user(body.email, body.password)
    if not user:
        raise HTTPException(401, "Nieprawidłowy email lub hasło")
    return {
        "token":    _make_token(user["id"]),
        "name":     user["name"],
        "email":    user["email"],
        "is_admin": user["is_admin"],
    }


@app.get("/auth/me")
def me(user: dict = Depends(get_auth_user)):
    return {k: v for k, v in user.items() if k != "password_hash"}


# ── admin user management ─────────────────────────────────────────────────────

class CreateUserIn(BaseModel):
    email: str
    name:  str
    password: str
    is_admin: bool = False


class SetPasswordIn(BaseModel):
    password: str


class SetActiveIn(BaseModel):
    active: bool


@app.get("/admin/users")
def admin_list_users(_: dict = Depends(require_admin)):
    return list_users()


@app.post("/admin/users")
def admin_create_user(body: CreateUserIn, _: dict = Depends(require_admin)):
    if len(body.password) < 6:
        raise HTTPException(400, "Hasło musi mieć min. 6 znaków")
    try:
        return create_user_db(body.email, body.name, body.password, body.is_admin)
    except Exception as e:
        raise HTTPException(409 if "unique" in str(e).lower() else 500, str(e))


@app.post("/admin/users/{uid}/set-password")
def admin_set_password(uid: int, body: SetPasswordIn, _: dict = Depends(require_admin)):
    if len(body.password) < 6:
        raise HTTPException(400, "Hasło musi mieć min. 6 znaków")
    reset_user_password(uid, body.password)
    return {"ok": True}


@app.post("/admin/users/{uid}/set-active")
def admin_set_active(uid: int, body: SetActiveIn, _: dict = Depends(require_admin)):
    set_user_status(uid, body.active)
    return {"ok": True}


# ── shared helpers ────────────────────────────────────────────────────────────

def norm_sn(s) -> str:
    if not s or str(s).strip() in ("", "nan", "None"):
        return ""
    s = str(s).strip()
    if s.endswith(".0"):        # int stored as float in Excel
        s = s[:-2]
    return re.sub(r"^0+", "", s).upper()


_PL = str.maketrans("ąćęłńóśźżĄĆĘŁŃÓŚŹŻ", "acelnoszzACELNOSZZ")

def _n(s: str) -> str:
    """Lowercase + strip Polish diacritics for fuzzy header matching."""
    return str(s).lower().translate(_PL)


def find_col(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
    """Case-insensitive, diacritics-insensitive column search."""
    for kw in keywords:
        nkw = _n(kw)
        for col in df.columns:
            if nkw in _n(col):
                return col
    return None


_SN_RE = re.compile(r'^[A-Z0-9]{3,15}$')

def _clean_sn_val(v: str) -> str:
    """Strip float suffix (.0) and normalize for SN pattern matching."""
    v = str(v).strip().upper()
    if v.endswith(".0"):
        v = v[:-2]
    return v

def find_sn_col_by_data(df: pd.DataFrame) -> Optional[str]:
    """
    Fallback: gdy żaden nagłówek nie pasuje do SN (np. pusta/scalona komórka w ODS),
    szukamy kolumny której wartości wyglądają jak numery seryjne (krótkie alfanumeryczne).
    Preferujemy kolumny 'Unnamed' zaraz po kolumnie z nazwą maszyny.
    ODS/Excel często eksportuje liczby z '.0' — strippujemy przed dopasowaniem.
    """
    unnamed_cols = [c for c in df.columns if str(c).startswith("Unnamed")]
    maszyna_idx = next(
        (i for i, c in enumerate(df.columns) if _n("maszyna") in _n(c) or _n("model") in _n(c)),
        -1,
    )
    if maszyna_idx >= 0:
        ordered = sorted(
            unnamed_cols,
            key=lambda c: abs(list(df.columns).index(c) - maszyna_idx - 1),
        )
    else:
        ordered = unnamed_cols

    for col in ordered:
        raw_vals = [str(v).strip() for v in df[col]
                    if str(v).strip() not in ("", "nan", "NaN", "NAN", "None", "NONE")]
        sample = [_clean_sn_val(v) for v in raw_vals[:50]]
        if not sample:          # kolumna całkowicie pusta
            continue
        hits = sum(1 for v in sample if _SN_RE.match(v))
        # Progi trafień (hits/sample):
        #  - 1 wpis  → 100% (unikamy pojedynczego fałszywego trafienia)
        #  - 2-5     → 50%  (dopuszcza 1 wiersz nagłówkowy + 1-4 prawdziwe SN-y,
        #                    typowe dla pliku z 1 urządzeniem + 2-nagłówkowy ODS)
        #  - 6+      → 35%  (dla plików z wierszami podsumowań / tytułowymi)
        if len(sample) == 1:
            threshold = 1.0
        elif len(sample) <= 5:
            threshold = 0.5
        else:
            threshold = 0.35
        if hits / len(sample) >= threshold:
            return col

    return None


def fmt_date(raw) -> str:
    """
    Parse a date value and return YYYY-MM-DD.
    Handles Polish DD.MM.YYYY format explicitly — avoids the dayfirst ambiguity bug
    in pd.to_datetime (which is 'not strict' and may swap day/month when day ≤ 12).
    """
    if raw is None or str(raw).strip() in ("", "nan", "NaT", "None"):
        return ""

    # If already a datetime-like object (e.g. from Excel date cell), use it directly
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")

    s = str(raw).strip()

    # 1. Polish/European format: DD.MM.YYYY or DD-MM-YYYY or DD/MM/YYYY
    #    Always treat first component as DAY, second as MONTH — no ambiguity.
    m = re.match(r'^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})$', s)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                from datetime import date as _date
                return _date(year, month, day).strftime("%Y-%m-%d")
            except ValueError:
                return f"{year}-{month:02d}"

    # 2. ISO format: YYYY-MM-DD (what Excel/pandas gives for date cells)
    m = re.match(r'^(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', s)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                from datetime import date as _date
                return _date(year, month, day).strftime("%Y-%m-%d")
            except ValueError:
                return f"{year}-{month:02d}"

    # 3. Already YYYY-MM-DD or YYYY-MM
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', s)
    if m:
        return m.group(1)
    m = re.match(r'^(\d{4}-\d{2})$', s)
    if m:
        return m.group(1)

    # 4. Pandas fallback (last resort — still uses dayfirst hint)
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="raise")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""


MONTH_RE = re.compile(
    r"\d{4}[-/\.]\d{2}"
    r"|\d{2}[-/\.]\d{4}"
    r"|sty|lut|mar|kwi|maj|cze|lip|sie|wrz|paź|paz|lis|gru"
    r"|jan|feb|apr|jun|jul|aug|sep|oct|nov|dec",
    re.IGNORECASE,
)


def detect_month_cols(df: pd.DataFrame) -> List[str]:
    return [c for c in df.columns if MONTH_RE.search(str(c))]


def detect_sep(data: bytes) -> str:
    """Sniff CSV separator from the first 4 KB (handles UTF-8 BOM)."""
    sample = data[:4096].decode("utf-8-sig", errors="ignore")
    return ";" if sample.count(";") > sample.count(",") else ","


def read_upload(data: bytes, filename: str) -> pd.DataFrame:
    name = filename.lower()
    if name.endswith(".csv"):
        sep = detect_sep(data)
        # Twoje pliki "Rozliczenie produkcyjne" mają wiersze podsumowań z >N kolumn
        # on_bad_lines='skip' pomija takie wiersze automatycznie.
        # Kodowanie: UTF-8-BOM → CP1250 → ISO-8859-2 (polskie excelowe eksporty)
        for enc in ("utf-8-sig", "cp1250", "iso-8859-2"):
            try:
                df = pd.read_csv(
                    io.BytesIO(data), dtype=str, encoding=enc, sep=sep,
                    on_bad_lines="skip",
                )
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Nie można odczytać pliku CSV (spróbuj zapisać jako UTF-8).")
    elif name.endswith(".ods"):
        # ODS (LibreOffice/OpenOffice) — wymaga odfpy
        df = pd.read_excel(io.BytesIO(data), dtype=str, engine="odf")
    else:
        # XLSX / XLS
        df = pd.read_excel(io.BytesIO(data), dtype=str)
    df = df.fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ── endpoints ─────────────────────────────────────────────────────────────────

@app.get("/status")
def db_status():
    """How many devices and which payment months are in the DB."""
    try:
        return get_status()
    except Exception as e:
        raise HTTPException(503, f"Błąd bazy danych: {e}")


@app.post("/import/production")
async def import_production(
    files:           List[UploadFile] = File(...),
    mode:            Annotated[str, Form()] = "append",
    device_type_tag: Annotated[str, Form()] = "",
):
    """
    Import one or more production files (zestawienie produkcji).
    mode=append  (default) → INSERT … ON CONFLICT DO NOTHING (skip existing SNs)
    mode=overwrite         → INSERT … ON CONFLICT DO UPDATE  (update existing SNs)
    device_type_tag        → if set (e.g. 'stare'), sets device_type_override for inserted SNs
    """
    total      = 0
    skipped    = 0
    all_new_sns: List[str] = []
    errors: List[str] = []

    for file in files:
        try:
            raw_data = await file.read()
            df = read_upload(raw_data, file.filename)

            _SN_KW = ["numer seryjny", "serial", "nr_ser", "nr ser", "sn",
                      "nr.ser", "nr. ser", "nr.seryjny", "nr seryjny",
                      "number", "device", "sn.", "s/n"]

            sn_col = find_col(df, _SN_KW)
            # Fallback: nagłówek może być pusty (scalona komórka w ODS/XLSX)
            if not sn_col:
                sn_col = find_sn_col_by_data(df)

            # Dla plików ODS: starsze pliki (2018-2019) mogą mieć wiersz tytułowy
            # w wierszu 0, a rzeczywiste nagłówki dopiero w wierszu 1 lub 2.
            # Próbujemy ponownie z różnymi wartościami parametru header.
            if not sn_col and file.filename.lower().endswith(".ods"):
                for hdr in (1, 2):
                    try:
                        df2 = pd.read_excel(
                            io.BytesIO(raw_data),
                            dtype=str, engine="odf", header=hdr,
                        )
                        df2 = df2.fillna("")
                        df2.columns = [str(c).strip() for c in df2.columns]
                        sn2 = find_col(df2, _SN_KW)
                        if not sn2:
                            sn2 = find_sn_col_by_data(df2)
                        if sn2:
                            df = df2
                            sn_col = sn2
                            break
                    except Exception:
                        pass

            if not sn_col:
                cols_preview = " | ".join(str(c) for c in df.columns[:15])
                # Dołącz próbkę wartości z kolumn Unnamed — pomaga debugować format
                unnamed_samples = []
                for c in df.columns:
                    if str(c).startswith("Unnamed"):
                        vals = [str(v).strip() for v in df[c]
                                if str(v).strip() not in ("", "nan")][:3]
                        if vals:
                            unnamed_samples.append(f"{c}: {vals}")
                hint = ("; próbki: " + ", ".join(unnamed_samples)) if unnamed_samples else ""
                errors.append(
                    f"{file.filename}: nie znaleziono kolumny z SN. "
                    f"Kolumny: [{cols_preview}]{hint}"
                )
                continue

            # ── column mapping ─────────────────────────────────────────────
            # "Rozliczenie produkcyjne" format (Twoje pliki):
            #   OPERATOR      = klient (CANIS, HHW, BEESWIFT…)   → firma
            #   Wdrożeniowiec = wdrożeniowiec (WIĘZIK, WÓJTOWICZ) → operator
            # Generic format:
            #   firma/company/klient/zamow → firma
            #   operator/handlowiec        → operator
            wdrozeniowiec_col = find_col(df, ["wdrożeniowiec", "wdrozeniowiec", "wdroż", "wdroz"])

            if wdrozeniowiec_col:
                # Twój format – OPERATOR = klient, Wdrożeniowiec = implementer
                firma_col    = find_col(df, ["operator"])
                operator_col = wdrozeniowiec_col
            else:
                firma_col    = find_col(df, ["firma", "company", "klient", "zamow"])
                operator_col = find_col(df, ["operator", "handlowiec"])

            maszyna_col = find_col(df, ["maszyna", "model", "typ", "machine"])
            date_col    = find_col(df, ["data produkcji", "data", "date"])

            records = []
            for _, row in df.iterrows():
                sn = norm_sn(row.get(sn_col, ""))
                if not sn:
                    continue
                records.append((
                    sn,
                    str(row.get(firma_col,    "")).strip() if firma_col    else "",
                    str(row.get(maszyna_col,  "")).strip() if maszyna_col  else "",
                    str(row.get(operator_col, "")).strip() if operator_col else "",
                    fmt_date(row.get(date_col, ""))         if date_col     else "",
                ))

            if records:
                # Deduplikacja po SN w obrębie jednego pliku (zachowaj ostatnie)
                deduped = list({r[0]: r for r in records}.values())
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        if mode == "overwrite":
                            execute_values(cur, """
                                INSERT INTO devices (sn, firma, maszyna, operator, prod_date)
                                VALUES %s
                                ON CONFLICT (sn) DO UPDATE SET
                                    firma      = EXCLUDED.firma,
                                    maszyna    = EXCLUDED.maszyna,
                                    operator   = EXCLUDED.operator,
                                    prod_date  = EXCLUDED.prod_date,
                                    updated_at = NOW()
                            """, deduped)
                            total += len(deduped)
                            all_new_sns.extend([r[0] for r in deduped])
                        else:  # append — skip existing SNs
                            rows = execute_values(cur, """
                                INSERT INTO devices (sn, firma, maszyna, operator, prod_date)
                                VALUES %s
                                ON CONFLICT (sn) DO NOTHING
                                RETURNING sn
                            """, deduped, fetch=True)
                            new_sns = [r[0] for r in rows]
                            total   += len(new_sns)
                            skipped += len(deduped) - len(new_sns)
                            all_new_sns.extend(new_sns)

        except Exception as e:
            errors.append(f"{file.filename}: {e}")

    # Jeśli podano device_type_tag (np. 'stare') — ustaw override dla nowych SN
    if device_type_tag and all_new_sns:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE devices SET device_type_override = %s WHERE sn = ANY(%s)",
                    (device_type_tag, all_new_sns),
                )

    # Zapisz sesję importu
    filenames_str = ", ".join(f.filename for f in files)
    create_import_session(
        import_type="production",
        filenames=filenames_str,
        mode=mode,
        device_type_tag=device_type_tag,
        records_added=total,
        records_skipped=skipped,
        sns=all_new_sns,
    )

    _invalidate_cache()
    return {"imported": total, "skipped": skipped, "mode": mode,
            "device_type_tag": device_type_tag, "errors": errors}


@app.post("/import/payments")
async def import_payments(
    file:       UploadFile = File(...),
    year_month: Annotated[Optional[str], Form()] = None,
    mode:       Annotated[str, Form()] = "append",
):
    """
    Obsługuje trzy formaty pliku płatności:

    1. Format IDS (surowy z ERP) — wykrywany automatycznie po kolumnie
       FS_DataWystawienia: każdy wiersz = jedna faktura, data wystawienia
       określa miesiąc płatności.

    2. Tabela przestawna — kolumny = miesiące, komórka = wartość płatności.

    3. Lista miesięczna — jeden miesiąc na raz; podaj year_month = 'YYYY-MM'.
    """
    df = read_upload(await file.read(), file.filename)

    # SN: tw_SN (IDS) lub generyczny
    sn_col = find_col(df, ["tw_sn", "numer seryjny", "serial", "nr_ser", "nr ser"])
    if not sn_col:
        raise HTTPException(422, "Brak kolumny z numerem seryjnym. Oczekiwano: tw_SN lub 'Numer seryjny'.")

    # Kolumna klienta: IDS używa adr_NazwaPelna
    cust_col = find_col(df, ["adr_nazwapelna", "adr_nazwa", "klient", "client", "customer", "nazwa"])

    # Data wystawienia faktury (format IDS)
    invoice_date_col = find_col(df, ["fs_datawystawienia", "datawystawienia", "data wystawienia", "fs_data"])

    month_cols = detect_month_cols(df)

    records: List[tuple] = []

    if invoice_date_col:
        # ── Format IDS: surowy eksport z ERP ─────────────────────────────────
        # Każdy wiersz = jedna faktura; data wystawienia → miesiąc płatności.
        # Kolumna ob_Ilosc wskazuje ile miesięcy obejmuje faktura (np. 12 dla
        # rocznego abonamentu) — generujemy tyle kolejnych rekordów miesięcznych.
        qty_col      = find_col(df, ["ob_ilosc", "ob_il", "ilosc", "quantity", "qty"])
        cat_col      = find_col(df, ["fs_kategoria", "fs_kat", "kategoria", "category"])
        # Data ostatniej spłaty — jeśli pusta, faktura nieopłacona → pomijamy
        pay_date_col = find_col(df, ["nzf_dataostatniejsplaty", "dataostatniejsplaty",
                                     "datazaplaty", "fs_datazaplaty", "data_zaplaty",
                                     "ostatniasplata", "ostatniazaplata"])
        # Kwota w walucie oryginalnej → fallback netto PLN (bez nzf_WartoscPierwotnaWaluta — złe dane)
        amount_col      = find_col(df, ["ob_cenawaluta", "ob_cenanetto",
                                        "ob_wartosc", "wartosc", "fs_wartosc", "amount", "kwota"])
        currency_col    = find_col(df, ["nzf_idwaluty", "idwaluty", "waluta", "currency"])
        # Kwota netto PLN i brutto PLN (dodatkowe kolumny)
        netto_col       = find_col(df, ["ob_cenanetto", "cenanetto", "netto", "net"])
        brutto_col      = find_col(df, ["ob_cenabrutto", "cenabrutto", "brutto", "gross"])

        skipped = 0
        skipped_unpaid = 0
        for _, row in df.iterrows():
            sn = norm_sn(row.get(sn_col, ""))
            if not sn:
                continue

            # ── WALIDACJA: data ostatniej spłaty musi być niepusta ──────────
            if pay_date_col:
                pay_date_val = str(row.get(pay_date_col, "")).strip()
                if pay_date_val in ("", "nan", "None", "NaT", "NaN"):
                    skipped_unpaid += 1
                    continue   # faktura wystawiona ale nieopłacona — pomijamy

            ym_full = fmt_date(row.get(invoice_date_col, ""))
            if not ym_full:
                skipped += 1
                continue
            ym = ym_full[:7]   # zawsze YYYY-MM — unikamy duplikatów z różnych dni
            customer = str(row.get(cust_col, "")).strip() if cust_col else ""

            # Liczba miesięcy z ob_Ilosc
            months_count = 1
            if qty_col:
                try:
                    q = float(str(row.get(qty_col, "1")).replace(",", "."))
                    if 1 <= q <= 120:   # rozsądny zakres 1-120 mies.
                        months_count = int(q)
                except (ValueError, TypeError):
                    pass

            # Fallback: parsuj "co 12 miesięcy" z FS_Kategoria
            if months_count == 1 and cat_col:
                cat_str = str(row.get(cat_col, ""))
                m = re.search(r'co\s+(\d+)\s+miesi', cat_str, re.IGNORECASE)
                if m:
                    months_count = int(m.group(1))

            def _parse_amt(col):
                if not col: return 0.0
                try:
                    raw = str(row.get(col, "0")).replace(",", ".").replace(" ", "").replace("\xa0", "")
                    v = float(raw)
                    return v if v > 0 else 0.0
                except (ValueError, TypeError):
                    return 0.0

            # Kwota w walucie oryginalnej — dzielimy proporcjonalnie na miesiące
            total_amount  = _parse_amt(amount_col)
            total_netto   = _parse_amt(netto_col)
            total_brutto  = _parse_amt(brutto_col)

            amount_per_month = round(total_amount / months_count, 2) if months_count > 1 else total_amount
            netto_per_month  = round(total_netto  / months_count, 2) if months_count > 1 else total_netto
            brutto_per_month = round(total_brutto / months_count, 2) if months_count > 1 else total_brutto

            currency = ""
            if currency_col:
                currency = str(row.get(currency_col, "")).strip()
                if currency in ("nan", "None"):
                    currency = ""

            # Generuj rekord dla każdego miesiąca objętego fakturą
            base_y, base_m = int(ym[:4]), int(ym[5:7])
            for i in range(months_count):
                offset = base_m - 1 + i
                rec_y = base_y + offset // 12
                rec_m = offset % 12 + 1
                records.append((sn, f"{rec_y}-{rec_m:02d}", customer,
                                 amount_per_month, currency,
                                 netto_per_month, brutto_per_month))

    elif month_cols:
        # ── Tabela przestawna ─────────────────────────────────────────────────
        for _, row in df.iterrows():
            sn = norm_sn(row.get(sn_col, ""))
            if not sn:
                continue
            customer = str(row.get(cust_col, "")).strip() if cust_col else ""
            for mc in month_cols:
                val = str(row.get(mc, "")).strip()
                if val and val not in ("0", "nan", "None", ""):
                    ym = fmt_date(mc) or str(mc).strip()[:7]
                    if ym:
                        # Kwota z wartości komórki (jeśli liczbowa)
                        try:
                            amt = round(float(val.replace(",", ".").replace(" ", "")), 2)
                        except (ValueError, TypeError):
                            amt = 0.0
                        records.append((sn, ym, customer, amt, ""))

    elif year_month:
        # ── Lista miesięczna ──────────────────────────────────────────────────
        paid_col = find_col(df, ["oplacony", "oplacone", "paid", "status"])
        amount_col_m = find_col(df, ["kwota", "wartosc", "amount"])
        for _, row in df.iterrows():
            sn = norm_sn(row.get(sn_col, ""))
            if not sn:
                continue
            customer = str(row.get(cust_col, "")).strip() if cust_col else ""
            if paid_col:
                val = str(row.get(paid_col, "")).strip().lower()
                if val in ("0", "nie", "no", "false", "", "nan"):
                    continue
            amt = 0.0
            if amount_col_m:
                try:
                    amt = round(float(str(row.get(amount_col_m, "0")).replace(",", ".").replace(" ", "")), 2)
                except (ValueError, TypeError):
                    pass
            records.append((sn, year_month, customer, amt, ""))

    else:
        raise HTTPException(
            422,
            "Nie rozpoznano formatu pliku. Oczekiwano kolumny FS_DataWystawienia "
            "(format IDS), kolumn miesięcy (tabela przestawna) lub podaj miesiąc ręcznie.",
        )

    if not records:
        return {"inserted": 0, "months": [], "message": "Brak rekordów do importu."}

    # Deduplikacja po (sn, year_month) — wiele faktur dla tego SN w tym samym miesiącu
    # zachowujemy pierwszą z niepustym klientem i wyższą kwotą
    deduped_pay: dict = {}
    for rec in records:
        sn, ym, customer, amount, currency = rec[0], rec[1], rec[2], rec[3], rec[4]
        netto   = rec[5] if len(rec) > 5 else 0.0
        brutto  = rec[6] if len(rec) > 6 else 0.0
        key = (sn, ym)
        if key not in deduped_pay:
            deduped_pay[key] = (sn, ym, customer, amount, currency, netto, brutto)
        else:
            ex = deduped_pay[key]
            deduped_pay[key] = (
                sn, ym,
                ex[2] if ex[2] else customer,
                ex[3] if ex[3] >= amount else amount,
                ex[4] if ex[4] else currency,
                ex[5] if ex[5] >= netto   else netto,
                ex[6] if ex[6] >= brutto  else brutto,
            )
    deduped_list = list(deduped_pay.values())

    pay_inserted = 0
    pay_skipped  = 0
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if mode == "overwrite":
                    execute_values(cur, """
                        INSERT INTO payments (sn, year_month, customer, amount, currency, amount_netto, amount_brutto)
                        VALUES %s
                        ON CONFLICT (sn, year_month) DO UPDATE SET
                            customer      = CASE WHEN EXCLUDED.customer <> '' THEN EXCLUDED.customer
                                                 ELSE payments.customer END,
                            amount        = CASE WHEN EXCLUDED.amount   > 0   THEN EXCLUDED.amount
                                                 ELSE payments.amount   END,
                            currency      = CASE WHEN EXCLUDED.currency <> '' THEN EXCLUDED.currency
                                                 ELSE payments.currency END,
                            amount_netto  = CASE WHEN EXCLUDED.amount_netto  > 0 THEN EXCLUDED.amount_netto
                                                 ELSE payments.amount_netto  END,
                            amount_brutto = CASE WHEN EXCLUDED.amount_brutto > 0 THEN EXCLUDED.amount_brutto
                                                 ELSE payments.amount_brutto END
                    """, deduped_list)
                    pay_inserted = len(deduped_list)
                else:  # append — skip existing (sn, year_month) pairs
                    rows = execute_values(cur, """
                        INSERT INTO payments (sn, year_month, customer, amount, currency, amount_netto, amount_brutto)
                        VALUES %s
                        ON CONFLICT (sn, year_month) DO NOTHING
                        RETURNING sn, year_month
                    """, deduped_list, fetch=True)
                    pay_inserted = len(rows)
                    pay_skipped  = len(deduped_list) - len(rows)
    except Exception as e:
        raise HTTPException(500, f"Błąd zapisu do bazy: {e}")

    months = sorted({r[1] for r in deduped_list})
    fmt = "IDS" if invoice_date_col else ("pivot" if month_cols else "monthly")

    skipped_unpaid_count = skipped_unpaid if invoice_date_col else 0

    # Zapisz sesję importu płatności
    if mode == "append":
        # rows zawiera (sn, year_month) dla faktycznie wstawionych rekordów
        pay_pairs = [(r[0], r[1]) for r in rows]
    else:
        pay_pairs = [(r[0], r[1]) for r in deduped_list]
    create_import_session(
        import_type="payments",
        filenames=file.filename,
        mode=mode,
        device_type_tag="",
        records_added=pay_inserted,
        records_skipped=pay_skipped,
        pay_pairs=pay_pairs,
    )

    _invalidate_cache()
    return {
        "inserted":          pay_inserted,
        "skipped":           pay_skipped,
        "mode":              mode,
        "months":            months,
        "format":            fmt,
        "duplicates_merged": len(records) - len(deduped_list),
        "skipped_unpaid":    skipped_unpaid_count,
        "pay_date_col":      pay_date_col if invoice_date_col else None,
        "amount_col":        amount_col   if invoice_date_col else None,
        "currency_col":      currency_col if invoice_date_col else None,
        "netto_col":         netto_col    if invoice_date_col else None,
        "brutto_col":        brutto_col   if invoice_date_col else None,
    }


@app.delete("/import/production")
def clear_production():
    """Remove all devices from the DB (use with caution)."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM devices")
            deleted = cur.rowcount
    _invalidate_cache()
    return {"deleted": deleted}


@app.delete("/import/payments")
def clear_payments(year_month: Optional[str] = None):
    """Remove payments — optionally only for a specific month."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            if year_month:
                cur.execute("DELETE FROM payments WHERE year_month = %s", (year_month,))
            else:
                cur.execute("DELETE FROM payments")
            deleted = cur.rowcount
    _invalidate_cache()
    return {"deleted": deleted, "year_month": year_month}


# ── in-memory cache for /analyze ─────────────────────────────────────────────
_analyze_cache: list | None = None

def _get_cached_analysis() -> list:
    """Return cached full result, or fetch from DB if cache is empty."""
    global _analyze_cache
    if _analyze_cache is None:
        _analyze_cache = get_analysis()
    return _analyze_cache

def _invalidate_cache() -> None:
    global _analyze_cache
    _analyze_cache = None


@app.get("/analyze")
def analyze(
    status:      Optional[str] = None,
    customer:    Optional[str] = None,
    operator:    Optional[str] = None,
    rep:         Optional[str] = None,
    device_type: Optional[str] = None,
    date_from:   Optional[str] = None,
    date_to:     Optional[str] = None,
):
    """Return joined result with optional filters. First call hits DB, subsequent calls use cache."""
    try:
        cached_rows = _get_cached_analysis()
    except Exception as e:
        raise HTTPException(503, f"Błąd bazy danych: {e}")

    # Overlay active suspensions (small DB query, not cached — time-sensitive)
    try:
        susp_map = get_all_suspensions_map()
    except Exception:
        susp_map = {"sns": {}, "firms": {}}

    susp_sns   = susp_map["sns"]
    susp_firms = susp_map["firms"]

    rows = []
    for r in cached_rows:
        row = dict(r)
        sn    = row.get("sn", "")
        firma = row.get("firma", "")
        is_suspended = sn in susp_sns or firma in susp_firms
        row["is_suspended"] = is_suspended
        rows.append(row)

    # Apply filters in Python on cached data (fast O(n) scan)
    if status == "suspended":
        rows = [r for r in rows if r["is_suspended"]]
    elif status:
        rows = [r for r in rows if r["status"] == status and not r["is_suspended"]]
    if customer:
        rows = [r for r in rows if r["customer"] == customer]
    if operator:
        rows = [r for r in rows if r["operator"] == operator]
    if rep:
        rows = [r for r in rows if rep in r["handlowcy"]]
    if device_type:
        rows = [r for r in rows if r["device_type"] == device_type]
    if date_from:
        rows = [r for r in rows if r["prod_date"] and r["prod_date"] >= date_from]
    if date_to:
        rows = [r for r in rows if not r["prod_date"] or r["prod_date"] <= date_to]

    suspended   = sum(1 for r in rows if r["is_suspended"])
    paid        = sum(1 for r in rows if r["status"] == "paid"   and not r["is_suspended"])
    unpaid      = sum(1 for r in rows if r["status"] == "unpaid" and not r["is_suspended"])
    only        = sum(1 for r in rows if r["status"] == "only"   and not r["is_suspended"])
    excluded    = sum(1 for r in rows if r["status"] == "excluded")
    oem         = sum(1 for r in rows if r["status"] == "oem")
    master_paid = sum(1 for r in rows if r["status"] == "paid" and r["device_type"] == "master" and not r["is_suspended"])

    return {
        "results": rows,
        "summary": {
            "total":       len(rows),
            "paid":        paid,
            "unpaid":      unpaid,
            "only":        only,
            "excluded":    excluded,
            "oem":         oem,
            "suspended":   suspended,
            "noBill":      oem + excluded,
            "masterPaid":  master_paid,
            "pct":         round(master_paid / (master_paid + unpaid) * 100) if (master_paid + unpaid) > 0 else 0,
        },
    }


# ── exclusions ────────────────────────────────────────────────────────────────

class ExclusionIn(BaseModel):
    firma: str
    reason: str = ""


@app.get("/exclusions")
def list_exclusions():
    try:
        return get_excluded_firms()
    except Exception as e:
        raise HTTPException(503, str(e))


@app.post("/exclusions")
def create_exclusion(body: ExclusionIn):
    try:
        add_excluded_firm(body.firma, body.reason)
        _invalidate_cache()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/exclusions")
def delete_exclusion(firma: str):
    try:
        remove_excluded_firm(firma)
        _invalidate_cache()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── sales reps ────────────────────────────────────────────────────────────────

class FirmAssign(BaseModel):
    firma: str


@app.get("/reps")
def list_reps():
    try:
        return get_reps()
    except Exception as e:
        raise HTTPException(503, str(e))


@app.get("/reps/firms")
def list_all_firms_for_reps():
    try:
        return {"firms": get_all_firms()}
    except Exception as e:
        raise HTTPException(503, str(e))


@app.post("/reps/{rep_id}/firms")
def add_firm_to_rep(rep_id: int, body: FirmAssign):
    try:
        assign_firm_to_rep(rep_id, body.firma)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/reps/{rep_id}/firms")
def remove_firm_from_rep_endpoint(rep_id: int, firma: str):
    try:
        remove_firm_from_rep(rep_id, firma)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── device type overrides ─────────────────────────────────────────────────────

class DeviceTypeIn(BaseModel):
    device_type: str            # 'master', 'oem', 'showroom', or '' (reset)
    showroom_until: Optional[str] = ""   # YYYY-MM, only used when device_type='showroom'


@app.patch("/devices/{sn}/type")
def override_device_type(sn: str, body: DeviceTypeIn):
    if body.device_type not in ("", "master", "oem", "showroom"):
        raise HTTPException(400, "device_type must be 'master', 'oem', 'showroom', or '' (reset)")
    if body.device_type == "showroom" and not body.showroom_until:
        raise HTTPException(400, "showroom_until (YYYY-MM) is required for showroom type")
    try:
        set_device_type_override(sn, body.device_type, body.showroom_until or "")
        _invalidate_cache()
        return {"ok": True, "sn": sn, "device_type": body.device_type,
                "showroom_until": body.showroom_until}
    except Exception as e:
        raise HTTPException(500, str(e))


class CommentIn(BaseModel):
    comment: str = ""


@app.patch("/devices/{sn}/comment")
def update_device_comment(sn: str, body: CommentIn):
    """Set or clear the free-text comment on a device."""
    try:
        set_device_comment(sn, body.comment)
        return {"ok": True, "sn": sn}
    except Exception as e:
        raise HTTPException(500, str(e))


class BulkTypeIn(BaseModel):
    sns:           List[str]
    device_type:   str            # 'master', 'oem', or '' (reset)
    showroom_until: Optional[str] = ""


@app.post("/devices/bulk-type")
def bulk_override_type(body: BulkTypeIn):
    """Change device_type_override for multiple SNs at once."""
    if body.device_type not in ("", "master", "oem"):
        raise HTTPException(400, "Bulk type must be 'master', 'oem', or '' (reset)")
    if not body.sns:
        raise HTTPException(400, "No SNs provided")
    try:
        count = bulk_set_device_type(body.sns, body.device_type)
        _invalidate_cache()
        return {"ok": True, "updated": count}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/devices/overrides")
def list_overrides():
    try:
        return get_type_overrides()
    except Exception as e:
        raise HTTPException(503, str(e))


@app.get("/payments/{sn}")
def payments_for_sn(sn: str):
    """Return all payment records for a single device (for history expand)."""
    try:
        return get_payments_for_sn(sn)
    except Exception as e:
        raise HTTPException(503, str(e))


@app.get("/health")
def health():
    return {"status": "ok"}


# ── sales rep management ──────────────────────────────────────────────────────

class SalesRepIn(BaseModel):
    name: str


@app.post("/sales-reps")
def create_sales_rep(body: SalesRepIn):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Nazwa handlowca nie może być pusta")
    try:
        new_id = add_sales_rep(name)
        return {"ok": True, "id": new_id, "name": name}
    except Exception as e:
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            raise HTTPException(409, f"Handlowiec '{name}' już istnieje")
        raise HTTPException(500, str(e))


@app.delete("/sales-reps/{rep_id}")
def delete_sales_rep(rep_id: int):
    try:
        remove_sales_rep(rep_id)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── monthly revenue ───────────────────────────────────────────────────────────

@app.get("/payments/monthly-revenue")
def monthly_revenue():
    try:
        return get_monthly_revenue()
    except Exception as e:
        raise HTTPException(503, str(e))


# ── firm config ───────────────────────────────────────────────────────────────

class FirmConfigIn(BaseModel):
    firm_type:       str   = "ids"
    cycle:           str   = ""
    expected_amount: float = 0.0
    currency:        str   = ""


@app.get("/firm-configs")
def list_firm_configs():
    try:
        return get_firm_configs()
    except Exception as e:
        raise HTTPException(503, str(e))


@app.put("/firm-configs/{firma}")
def set_firm_config(firma: str, body: FirmConfigIn):
    if body.firm_type not in VALID_FIRM_TYPES:
        raise HTTPException(400, f"firm_type musi być jednym z: {', '.join(VALID_FIRM_TYPES)}")
    if body.cycle not in VALID_CYCLES:
        raise HTTPException(400, f"cycle musi być jednym z: {', '.join(VALID_CYCLES)}")
    try:
        upsert_firm_config(firma, body.firm_type, body.cycle, body.expected_amount, body.currency)
        _invalidate_cache()
        return {"ok": True, "firma": firma}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/firm-configs/{firma}")
def remove_firm_config(firma: str):
    try:
        delete_firm_config(firma)
        _invalidate_cache()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── firm export / import ──────────────────────────────────────────────────────

@app.get("/firms/export")
def export_firms_excel():
    """Generate Excel file with all firms, types, cycles, amounts, reps (max 2)."""
    try:
        rows = get_firms_for_export()
    except Exception as e:
        raise HTTPException(503, str(e))

    CYCLE_PL = {"monthly": "miesięczny", "quarterly": "kwartalny",
                "annual": "roczny", "once": "jednorazowy", "": ""}
    TYPE_PL  = {"ids": "IDS", "licencja": "Licencja", "oem": "OEM", "inne": "Inne"}

    data = []
    for r in rows:
        reps = r.get("rep_names") or []
        data.append({
            "Firma":              r["firma"],
            "Typ firmy":          TYPE_PL.get(r["firm_type"], r["firm_type"]),
            "Cykl płatności":     CYCLE_PL.get(r["cycle"], r["cycle"]),
            "Oczekiwana kwota":   float(r["expected_amount"]) if r["expected_amount"] else 0,
            "Waluta":             r["currency"] or "",
            "Handlowiec 1":       reps[0] if len(reps) > 0 else "",
            "Handlowiec 2":       reps[1] if len(reps) > 1 else "",
        })

    df = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Firmy")
        ws = writer.sheets["Firmy"]
        # Column widths
        for i, w in enumerate([40, 12, 16, 18, 8, 28, 28], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        # Validation note in a second sheet
        info = writer.book.create_sheet("Instrukcja")
        notes = [
            ["Kolumna",           "Dozwolone wartości"],
            ["Typ firmy",         "IDS | Licencja | OEM | Inne"],
            ["Cykl płatności",    "miesięczny | kwartalny | roczny | jednorazowy | (puste)"],
            ["Oczekiwana kwota",  "liczba dziesiętna np. 120.00"],
            ["Waluta",            "PLN | EUR | USD | (puste)"],
            ["Handlowiec 1/2",    "Dokładna nazwa handlowca z systemu (wielkość liter ma znaczenie)"],
        ]
        for row in notes:
            info.append(row)
        info.column_dimensions["A"].width = 22
        info.column_dimensions["B"].width = 55

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=firmy_konfiguracja.xlsx"},
    )


class FirmsImportIn(BaseModel):
    mode: str = "supplement"   # supplement | overwrite


@app.post("/firms/import")
async def import_firms_excel(
    file: UploadFile = File(...),
    mode: Annotated[Optional[str], Form()] = "supplement",
):
    if mode not in ("supplement", "overwrite"):
        raise HTTPException(400, "mode musi być 'supplement' lub 'overwrite'")

    raw = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(raw), dtype=str).fillna("")
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        raise HTTPException(422, f"Nie można odczytać pliku Excel: {e}")

    TYPE_MAP = {"ids": "ids", "licencja": "licencja", "oem": "oem", "inne": "inne",
                "IDS": "ids", "Licencja": "licencja", "OEM": "oem", "Inne": "inne"}
    CYCLE_MAP = {
        "miesięczny": "monthly", "monthly": "monthly",
        "kwartalny":  "quarterly", "quarterly": "quarterly",
        "roczny":     "annual",   "annual": "annual",
        "jednorazowy":"once",     "once": "once",
        "": "",
    }

    # Map column headers flexibly
    col = lambda *keys: next((c for k in keys for c in df.columns if k.lower() in c.lower()), None)
    firma_col    = col("firma",    "company")
    type_col     = col("typ",      "type")
    cycle_col    = col("cykl",     "cycle")
    amount_col   = col("kwota",    "amount")
    currency_col = col("waluta",   "currency")
    rep1_col     = col("handlowiec 1", "rep1", "handlowiec1")
    rep2_col     = col("handlowiec 2", "rep2", "handlowiec2")

    if not firma_col:
        raise HTTPException(422, "Brak kolumny 'Firma' w pliku")

    rows = []
    for _, row in df.iterrows():
        firma = str(row.get(firma_col, "")).strip()
        if not firma:
            continue
        rows.append({
            "firma":           firma,
            "firm_type":       TYPE_MAP.get(str(row.get(type_col, "")).strip(), "ids") if type_col else "ids",
            "cycle":           CYCLE_MAP.get(str(row.get(cycle_col, "")).strip(), "") if cycle_col else "",
            "expected_amount": str(row.get(amount_col, "0")).strip() if amount_col else "0",
            "currency":        str(row.get(currency_col, "")).strip().upper() if currency_col else "",
            "rep1":            str(row.get(rep1_col, "")).strip() if rep1_col else "",
            "rep2":            str(row.get(rep2_col, "")).strip() if rep2_col else "",
        })

    try:
        result = import_firms_table(rows, mode)
    except Exception as e:
        raise HTTPException(500, str(e))

    return {
        "ok":             True,
        "mode":           mode,
        "firms_processed": len(rows),
        **result,
    }


# ── revenue analytics ─────────────────────────────────────────────────────────

@app.get("/revenue")
def get_revenue_analytics():
    """
    Aggregated revenue data for the Wyliczenia tab.
    All amounts: amount_netto / amount_brutto always in PLN;
                 'eur' = SUM(amount) WHERE currency='EUR'.
    """
    cur_year = str(datetime.now().year)

    with get_conn() as conn:
        with conn.cursor() as cur:

            # 1. Monthly trend: PLN netto + EUR + brutto
            cur.execute("""
                SELECT year_month,
                       COALESCE(SUM(amount_netto),  0)::float  AS netto,
                       COALESCE(SUM(amount_brutto), 0)::float  AS brutto,
                       COALESCE(SUM(CASE WHEN currency='EUR' THEN amount ELSE 0 END), 0)::float AS eur
                FROM payments
                WHERE amount_netto > 0 OR amount_brutto > 0 OR amount > 0
                GROUP BY year_month
                ORDER BY year_month
            """)
            cols = ['month', 'netto', 'brutto', 'eur']
            monthly = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 2. Annual totals
            cur.execute("""
                SELECT LEFT(year_month, 4)                                             AS year,
                       COALESCE(SUM(amount_netto),  0)::float                          AS netto,
                       COALESCE(SUM(amount_brutto), 0)::float                          AS brutto,
                       COALESCE(SUM(CASE WHEN currency='EUR' THEN amount ELSE 0 END), 0)::float AS eur,
                       COUNT(DISTINCT sn)                                              AS devices,
                       COUNT(DISTINCT customer)                                        AS customers
                FROM payments
                WHERE amount_netto > 0 OR amount_brutto > 0 OR amount > 0
                GROUP BY LEFT(year_month, 4)
                ORDER BY year
            """)
            cols = ['year', 'netto', 'brutto', 'eur', 'devices', 'customers']
            annual = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 3. Top 50 customers by PLN netto
            cur.execute("""
                SELECT customer,
                       COALESCE(SUM(amount_netto),  0)::float                          AS netto,
                       COALESCE(SUM(amount_brutto), 0)::float                          AS brutto,
                       COALESCE(SUM(CASE WHEN currency='EUR' THEN amount ELSE 0 END), 0)::float AS eur,
                       COUNT(DISTINCT sn)                                              AS devices,
                       COUNT(*)                                                         AS payment_count
                FROM payments
                WHERE customer IS NOT NULL AND customer <> ''
                  AND (amount_netto > 0 OR amount_brutto > 0 OR amount > 0)
                GROUP BY customer
                ORDER BY netto DESC
                LIMIT 50
            """)
            cols = ['customer', 'netto', 'brutto', 'eur', 'devices', 'payment_count']
            customers = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 4. Revenue by sales rep (via firm_reps → devices)
            cur.execute("""
                SELECT sr.name                                                          AS rep,
                       COALESCE(SUM(p.amount_netto),  0)::float                        AS netto,
                       COALESCE(SUM(p.amount_brutto), 0)::float                        AS brutto,
                       COALESCE(SUM(CASE WHEN p.currency='EUR' THEN p.amount ELSE 0 END), 0)::float AS eur,
                       COUNT(DISTINCT p.sn)                                            AS devices
                FROM payments p
                JOIN devices d  ON d.sn     = p.sn
                JOIN firm_reps fr ON fr.firma = d.firma
                JOIN sales_reps sr ON sr.id  = fr.rep_id
                WHERE p.amount_netto > 0 OR p.amount_brutto > 0 OR p.amount > 0
                GROUP BY sr.name
                ORDER BY netto DESC
            """)
            cols = ['rep', 'netto', 'brutto', 'eur', 'devices']
            by_rep = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 5. Revenue by firm type
            cur.execute("""
                SELECT COALESCE(NULLIF(fc.firm_type,''), 'brak')                       AS firm_type,
                       COALESCE(SUM(p.amount_netto),  0)::float                        AS netto,
                       COALESCE(SUM(p.amount_brutto), 0)::float                        AS brutto,
                       COUNT(DISTINCT p.sn)                                            AS devices
                FROM payments p
                JOIN devices d ON d.sn = p.sn
                LEFT JOIN firm_config fc ON fc.firma = d.firma
                WHERE p.amount_netto > 0 OR p.amount_brutto > 0 OR p.amount > 0
                GROUP BY COALESCE(NULLIF(fc.firm_type,''), 'brak')
                ORDER BY netto DESC
            """)
            cols = ['type', 'netto', 'brutto', 'devices']
            by_type = [dict(zip(cols, r)) for r in cur.fetchall()]

            # 6. KPI: all-time totals
            cur.execute("""
                SELECT COALESCE(SUM(amount_netto),  0)::float  AS total_netto,
                       COALESCE(SUM(amount_brutto), 0)::float  AS total_brutto,
                       COALESCE(SUM(CASE WHEN currency='EUR' THEN amount ELSE 0 END), 0)::float AS total_eur,
                       COUNT(DISTINCT sn)                       AS paying_devices,
                       COUNT(DISTINCT customer)                 AS paying_customers
                FROM payments
                WHERE amount_netto > 0 OR amount_brutto > 0 OR amount > 0
            """)
            r = cur.fetchone()
            kpi_all = {
                'total_netto':       r[0], 'total_brutto':    r[1],
                'total_eur':         r[2], 'paying_devices':  r[3],
                'paying_customers':  r[4],
            }

            # 7. KPI: YTD (current year)
            cur.execute("""
                SELECT COALESCE(SUM(amount_netto),  0)::float  AS netto,
                       COALESCE(SUM(amount_brutto), 0)::float  AS brutto,
                       COALESCE(SUM(CASE WHEN currency='EUR' THEN amount ELSE 0 END), 0)::float AS eur,
                       COUNT(DISTINCT sn)                       AS devices,
                       COUNT(DISTINCT customer)                 AS customers
                FROM payments
                WHERE LEFT(year_month, 4) = %s
                  AND (amount_netto > 0 OR amount_brutto > 0 OR amount > 0)
            """, (cur_year,))
            r = cur.fetchone()
            kpi_ytd = {
                'year': cur_year, 'netto': r[0], 'brutto': r[1],
                'eur': r[2], 'devices': r[3], 'customers': r[4],
            }

    return {
        "monthly":   monthly,
        "annual":    annual,
        "customers": customers,
        "by_rep":    by_rep,
        "by_type":   by_type,
        "kpi_all":   kpi_all,
        "kpi_ytd":   kpi_ytd,
    }

# ── production seasonality ────────────────────────────────────────────────────

@app.get("/production/seasonality")
def production_seasonality(
    device_type: str = "master",   # master | oem | all
    model:       str = "",         # substring filter on maszyna (case-insensitive)
):
    """
    Month-by-month production counts per year for seasonality analysis.
    Uses the analysis cache — same device-type logic as the main report.
    Returns counts for months 1-12 per year.
    """
    try:
        rows = _get_cached_analysis()
    except Exception as e:
        raise HTTPException(503, str(e))

    # Filter by device type
    if device_type == "master":
        rows = [r for r in rows if r.get("device_type") == "master"]
    elif device_type == "oem":
        rows = [r for r in rows if r.get("status") == "oem"]
    elif device_type == "paid":
        rows = [r for r in rows if r.get("status") == "paid"]
    elif device_type == "unpaid":
        rows = [r for r in rows if r.get("status") == "unpaid"]
    # "all" → no type filter

    # Optional model filter
    if model:
        model_lc = model.lower()
        rows = [r for r in rows if model_lc in (r.get("maszyna") or "").lower()]

    # Group by year and calendar month (1-12)
    by_year: dict[str, dict[str, int]] = {}
    for r in rows:
        pd = r.get("prod_date") or ""
        if len(pd) < 7:
            continue
        year  = pd[:4]
        month = pd[5:7]
        if year not in by_year:
            by_year[year] = {}
        by_year[year][month] = by_year[year].get(month, 0) + 1

    years = sorted(by_year.keys())
    months_labels = ["Sty","Lut","Mar","Kwi","Maj","Cze",
                     "Lip","Sie","Wrz","Paź","Lis","Gru"]
    series = [
        {
            "year":   yr,
            "data":   [by_year[yr].get(f"{m:02d}", 0) for m in range(1, 13)],
            "total":  sum(by_year[yr].values()),
        }
        for yr in years
    ]

    # Also return list of unique models for the filter dropdown
    all_rows_for_models = _get_cached_analysis()
    models = sorted({
        r.get("maszyna", "").strip()
        for r in all_rows_for_models
        if r.get("maszyna", "").strip()
    })

    return {
        "years":         years,
        "months_labels": months_labels,
        "series":        series,
        "models":        models,
    }


# ── license fees ─────────────────────────────────────────────────────────────

class LicenseFeeIn(BaseModel):
    firma:     str
    amount:    float
    currency:  str   = "PLN"
    date_from: str   = ""
    date_to:   str   = ""
    note:      str   = ""


@app.get("/license-fees")
def list_license_fees():
    try:
        return get_license_fees()
    except Exception as e:
        raise HTTPException(503, str(e))


@app.post("/license-fees")
def create_license_fee(body: LicenseFeeIn):
    if not body.firma.strip():
        raise HTTPException(400, "Pole 'firma' jest wymagane")
    if body.amount <= 0:
        raise HTTPException(400, "Kwota musi być większa od 0")
    try:
        row = upsert_license_fee(
            body.firma.strip(), body.amount, body.currency.strip().upper(),
            body.date_from.strip(), body.date_to.strip(), body.note.strip()
        )
        return {"ok": True, "row": row}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.put("/license-fees/{fee_id}")
def update_license_fee(fee_id: int, body: LicenseFeeIn):
    if not body.firma.strip():
        raise HTTPException(400, "Pole 'firma' jest wymagane")
    if body.amount <= 0:
        raise HTTPException(400, "Kwota musi być większa od 0")
    try:
        row = upsert_license_fee(
            body.firma.strip(), body.amount, body.currency.strip().upper(),
            body.date_from.strip(), body.date_to.strip(), body.note.strip(),
            fee_id=fee_id,
        )
        return {"ok": True, "row": row}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/license-fees/{fee_id}")
def remove_license_fee(fee_id: int):
    try:
        delete_license_fee(fee_id)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── firms stats ───────────────────────────────────────────────────────────────

@app.get("/firms/stats")
def firms_stats():
    """Lista firm z liczbą urządzeń (dla podglądu stanu bazy)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT firma, COUNT(*) AS devices
                    FROM devices
                    WHERE firma <> ''
                    GROUP BY firma
                    ORDER BY firma
                """)
                rows = cur.fetchall()
        return {"firms": [{"firma": r[0], "devices": r[1]} for r in rows]}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── suspensions ───────────────────────────────────────────────────────────────

class SuspensionIn(BaseModel):
    date_from: str
    date_to:   str
    note:      str = ""


@app.get("/suspensions/active")
def suspensions_active():
    try:
        return get_active_suspensions()
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/devices/{sn}/suspensions")
def list_device_suspensions(sn: str):
    try:
        return {"suspensions": get_device_suspensions(sn)}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/devices/{sn}/suspensions")
def add_device_susp(sn: str, body: SuspensionIn):
    if body.date_from > body.date_to:
        raise HTTPException(400, "date_from nie może być późniejsza niż date_to")
    try:
        new_id = add_device_suspension(sn, body.date_from, body.date_to, body.note)
        return {"ok": True, "id": new_id}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/suspensions/device/{susp_id}")
def del_device_susp(susp_id: int):
    try:
        ok = delete_device_suspension(susp_id)
        if not ok:
            raise HTTPException(404, "Nie znaleziono zawieszenia")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/firms/{firma}/suspensions")
def list_firm_suspensions(firma: str):
    try:
        return {"suspensions": get_firm_suspensions(firma)}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/firms/{firma}/suspensions")
def add_firm_susp(firma: str, body: SuspensionIn):
    if body.date_from > body.date_to:
        raise HTTPException(400, "date_from nie może być późniejsza niż date_to")
    try:
        new_id = add_firm_suspension(firma, body.date_from, body.date_to, body.note)
        _invalidate_cache()
        return {"ok": True, "id": new_id}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/suspensions/firm/{susp_id}")
def del_firm_susp(susp_id: int):
    try:
        ok = delete_firm_suspension(susp_id)
        if not ok:
            raise HTTPException(404, "Nie znaleziono zawieszenia")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── firm-rep export / import ───────────────────────────────────────────────────

@app.get("/firms/reps/export")
def export_firm_reps():
    """Pobierz listę firm z handlowcami jako plik Excel."""
    try:
        rows = get_firms_with_reps()
    except Exception as e:
        raise HTTPException(500, str(e))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firmy_Handlowcy"
    ws.append(["firma", "handlowiec_1", "handlowiec_2"])
    for r in rows:
        ws.append([r["firma"], r["handlowiec_1"], r["handlowiec_2"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="firmy_handlowcy.xlsx"'}
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/firms/reps/import")
async def import_firm_reps_endpoint(file: UploadFile = File(...)):
    """Importuj przypisania handlowców z pliku Excel (tryb nadpisz)."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Nieprawidłowy plik Excel")
    ws = wb.active
    rows = []
    headers_done = False
    for row in ws.iter_rows(values_only=True):
        if not headers_done:
            headers_done = True
            continue
        if not row or not row[0]:
            continue
        rows.append({
            "firma":        str(row[0]).strip() if row[0] else "",
            "handlowiec_1": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "handlowiec_2": str(row[2]).strip() if len(row) > 2 and row[2] else "",
        })
    if not rows:
        raise HTTPException(400, "Plik nie zawiera danych")
    try:
        result = import_firm_reps(rows)
        _invalidate_cache()
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/firms/unassigned")
def firms_unassigned():
    """Firmy bez przypisanego handlowca."""
    try:
        return {"firms": get_firms_without_reps()}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── rep bonus check ────────────────────────────────────────────────────────────

@app.get("/reps/first-ids-check")
def rep_first_ids_check(
    rep_id:          int,
    first_pay_from:  str,
    first_pay_to:    str,
    window_months:   int = 12,
):
    """
    Pierwsze IDS — Handlowiec.
    Urządzenia handlowca, których PIERWSZA płatność mieści się w przedziale
    [first_pay_from, first_pay_to]. Okno: window_months od daty pierwszej płatności.
    """
    import re
    ym_re = r"^\d{4}-\d{2}$"
    if not re.match(ym_re, first_pay_from) or not re.match(ym_re, first_pay_to):
        raise HTTPException(400, "Daty muszą być w formacie YYYY-MM")
    if first_pay_from > first_pay_to:
        raise HTTPException(400, "first_pay_from nie może być późniejsza niż first_pay_to")
    if not 1 <= window_months <= 36:
        raise HTTPException(400, "window_months musi być między 1 a 36")
    try:
        return get_rep_first_ids_check(rep_id, first_pay_from, first_pay_to, window_months)
    except Exception as e:
        raise HTTPException(500, str(e))


# ── import sessions ───────────────────────────────────────────────────────────

@app.get("/import/sessions")
def list_import_sessions():
    """Historia sesji importu (50 ostatnich)."""
    try:
        sessions = get_import_sessions(50)
        # Serialize datetime to string
        for s in sessions:
            if s.get("created_at"):
                s["created_at"] = s["created_at"].strftime("%Y-%m-%d %H:%M")
        return {"sessions": sessions}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/import/sessions/{session_id}")
def delete_import_session(session_id: int):
    """Cofnij import: usuwa urządzenia/płatności z danej sesji."""
    try:
        result = undo_import_session(session_id)
        if "error" in result:
            raise HTTPException(404, result["error"])
        _invalidate_cache()
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── bulk device update ─────────────────────────────────────────────────────────

class BulkUpdateIn(BaseModel):
    sns:         List[str]
    firma:       Optional[str] = None
    operator:    Optional[str] = None
    device_type: Optional[str] = None


@app.patch("/devices/bulk")
def bulk_update_devices_endpoint(body: BulkUpdateIn, _=Depends(require_admin)):
    """Grupowa edycja wielu urządzeń naraz (firma / operator / device_type_override)."""
    if not body.sns:
        raise HTTPException(400, "Brak listy SN")
    valid_types = {"", "master", "slave", "oem", "showroom", "stare"}
    if body.device_type is not None and body.device_type not in valid_types:
        raise HTTPException(400, f"Nieznany typ: {body.device_type}")
    try:
        updated = bulk_update_devices(
            sns=body.sns,
            firma=body.firma,
            operator=body.operator,
            device_type=body.device_type,
        )
        _invalidate_cache()
        return {"updated": updated}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── company merge ─────────────────────────────────────────────────────────────

class MergeFirmsIn(BaseModel):
    source: str   # firm name that will be absorbed / renamed away
    target: str   # firm name that remains after merge


@app.get("/firms/merges")
def list_merges():
    """Historia scaleń firm."""
    try:
        return {"merges": get_merge_history()}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/firms/merge")
def merge_firms_endpoint(body: MergeFirmsIn):
    source = body.source.strip()
    target = body.target.strip()
    if not source or not target:
        raise HTTPException(400, "Pola 'source' i 'target' są wymagane")
    if source == target:
        raise HTTPException(400, "Źródło i cel są identyczne")
    try:
        counts = merge_firms(source, target)
        _invalidate_cache()
        return {"ok": True, "source": source, "target": target, "affected": counts}
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, str(e))
